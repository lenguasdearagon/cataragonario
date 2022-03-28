import os
import sys

from django.core.exceptions import ValidationError
from django.core.management.base import BaseCommand, CommandError
from linguatec_lexicon.models import DiatopicVariation, Entry, GramaticalCategory, Lexicon, Region, Word
from linguatec_lexicon.validators import validate_balanced_parenthesis
from openpyxl import load_workbook


class Command(BaseCommand):
    def add_arguments(self, parser):
        parser.add_argument('input_file', type=str)
        parser.add_argument('--drop', action='store_true', dest='drop', help="Drop existing data before import.")
        parser.add_argument(
            '--extract-regions', action='store_true', dest='extract_regions',
            help="Extract regions from input file.",
        )

    def handle(self, *args, **options):
        self.input_file = options['input_file']
        self.verbosity = options['verbosity']

        self.validate_input_file()
        self.lexicon = Lexicon.objects.get(name='castellano-catalán')

        if options['drop']:
            self.drop_existing_data()

        if options['extract_regions']:
            self.extract_regions_from_spreadsheet()
            sys.exit(0)

        self.populate_models()
        self.stdout.write("ES: {}   CAT: {}     FRANJA: {}".format(self.es, self.cat, self.aralan))
        self.stdout.write("rows: {}     errors: {}".format(self.rows, self.rows_errors))
        self.stdout.write("-- THE END! --")

    def validate_input_file(self):
        _, file_extension = os.path.splitext(self.input_file)
        if file_extension.lower() != '.xlsx':
            raise CommandError(
                'Unexpected filetype "{}". Should be an Excel document (XLSX)'.format(file_extension))

        if not os.path.exists(self.input_file):
            raise CommandError(
                'File "{}" not found.'.format(self.input_file)
            )

    def drop_existing_data(self):
        self.lexicon.words.all().delete()
        Entry.objects.all().delete()

    def populate_models(self):
        worksheets = self.load_worksheets()
        self.es = 0
        self.cat = 0
        self.aralan = 0
        self.rows = 0
        self.rows_errors = 0
        for ws in worksheets:
            for i, row in enumerate(ws.values):
                try:
                    row = self.clean_row(i + 1, row, ws.title)
                except EmptyRow:
                    continue
                except ValidationError as e:
                    # TODO(@slamora): don't save any value if there are errors
                    self.rows_errors += 1
                    continue

                self.rows += 1
                self.save_row(row)

    def extract_regions_from_spreadsheet(self):
        # TODO(@slamora) extract from all worksheets
        ws = self.load_first_worksheet()

        regions = []
        for i, row in enumerate(ws.values):
            # skip first row because contains headers
            # skip empty rows
            if i == 0 or not any(row):
                continue

            try:
                row = RowEntry(row, line_number=i + 1)
                row.clean()
            except ValidationError:
                import pprint
                msg = pprint.pformat(row.errors)
                self.stderr.write(msg)
                # raise

            regions += row.regions

        regions_grouped = {}
        for region, cities in regions:
            if region not in regions_grouped:
                regions_grouped[region] = set()
            regions_grouped[region].update(set(cities))

        self.stdout.write(pprint.pformat(regions_grouped))

        return regions_grouped

    def load_first_worksheet(self):
        wb = load_workbook(filename=self.input_file, read_only=True)

        if len(wb.worksheets) > 1:
            self.stderr.write("WARNING: only data of first worksheet will be imported.")

        return wb.worksheets[0]

    def load_worksheets(self):
        wb = load_workbook(filename=self.input_file, read_only=True)
        return wb.worksheets

    def clean_row(self, line_number, row, worksheet):
        if line_number == 1:
            raise EmptyRow('Skip header')    # skip first row because contains headers

        try:
            row = RowEntry(row, line_number=line_number, worksheet=worksheet)
            row.clean()
        except EmptyRow:
            raise
        except ValidationError:
            for error in row.errors:
                self.stderr.write(
                    "{:>8}.{:<4}: {:<15} {:<2} {:<10}".format(
                        worksheet, line_number, error["word"], error["column"], error["message"])
                )
            raise
        # handle another kind of errors
        except Exception as e:
            self.stderr.write(
                "{:>8}.{:<4}: {:<15} {:<2} {:<10}".format(
                    worksheet, line_number, error["word"], "", str(e))
            )

        return row

    def save_row(self, row):
            gramcats = row.gramcats
            for es_term in row.es:
                word, created = Word.objects.get_or_create(lexicon=self.lexicon, term=es_term)
                if created: self.es += 1

                # create entries of normalized catalan
                for cat_term in row.cat:
                    entry, cat_created = Entry.objects.get_or_create(
                        word=word, translation=cat_term, variation__isnull=True)
                    if cat_created:
                        entry.gramcats.set(gramcats)
                        self.cat += 1

                # create entries of dialectal catalan
                # DiatopicVariation == Cities | Valleys
                # Region == County
                for variation in row.variations:
                    entry, variation_created = Entry.objects.get_or_create(
                        word=word, translation=row.term, variation=variation)

                    if variation_created:
                        entry.gramcats.set(gramcats)
                        self.aralan += 1
                    elif not cat_created and not created:
                        # possible duplicate because word.term cat entry & variation entry already exists
                        msg = "Possible duplicated row (unique-entry): {} {}".format(row.term, cat_term)
                        self.stderr.write(
                           "{:>8}.{:<4}: {:<15} {:<2} {:<10}".format(
                                row.worksheet, row.line_number, word.term, "", msg)
                        )


def extract_regions(value):
    if value is None:
        return [('franja', ['general'])]

    value = value.strip().lower()
    validate_balanced_parenthesis(value)

    comma_position = value.find(",")
    regions = []
    if comma_position == -1:
        extracted_region = split_region_and_variants(value)
    else:
        parenthesis_beg_position = value.find("(")
        parenthesis_end_position = value.find(")")
        if comma_position > parenthesis_beg_position and comma_position < parenthesis_end_position:
            extracted_region = split_region_and_variants(value)
            next_comma_position = value.find(",", parenthesis_end_position)
            if next_comma_position == -1:
                raw = ''
            else:
                raw = value[next_comma_position + 1:]
        else:
            region_left = value[:comma_position]
            extracted_region = split_region_and_variants(region_left)
            raw = value[comma_position + 1:]

        if raw:
            regions += extract_regions(raw)

    regions = [extracted_region] + regions

    return regions


def split_region_and_variants(value):
    parenthesis_beg_position = value.find("(")
    parenthesis_end_position = value.find(")")

    if parenthesis_beg_position == -1:
        region = value
        variants = []
    else:
        region = value[:parenthesis_beg_position]
        variants = value[parenthesis_beg_position + 1:parenthesis_end_position]
        variants = split_and_strip(variants)

    return (region.strip(), variants)


def extract_variants(value):
    z = split_and_strip(value)
    print("VARIANT: ", z)
    return z


def split_and_strip(value):
    if value is None:
        raise TypeError("argument must be a string, not 'None'")
    return [item.strip() for item in value.split(',')]


class EmptyRow(Exception):
    pass


class RowEntry:
    fields = ['term', 'gramcats', 'regions', 'cat', 'es']
    default_error_messages = {
        'required': 'Required value (this column cannot be empty)',
    }

    def __init__(self, row, line_number, worksheet=None) -> None:
        self.row = row
        self.line_number = line_number
        self.worksheet = worksheet

    def clean(self):
        if not any(self.row):
            raise EmptyRow()
        self.errors = []
        self.cleaned_data = {}
        for i, fieldname in enumerate(self.fields):
            value = self.row[i]
            clean_method = getattr(self, "clean_{}".format(fieldname))
            self.cleaned_data[fieldname] = clean_method(value)
            setattr(self, fieldname, self.cleaned_data[fieldname])

        if self.errors:
            raise ValidationError("Errors on line: {}".format(self.line_number))

        return self.cleaned_data

    def clean_term(self, value):
        return value.strip()

    def clean_gramcats(self, value):
        if value is None:
            self.add_error("C", self.default_error_messages['required'])
            return []

        # TODO(@slamora) optimize queries
        gramcats = []
        for abbr in split_and_strip(value):
            try:
                gramcats.append(GramaticalCategory.objects.get(abbreviation=abbr))
            except GramaticalCategory.DoesNotExist:
                self.add_error("B", "unkown gramatical category '{}'".format(abbr))

        return gramcats

    def clean_regions(self, value):
        try:
            regions = extract_regions(value)
        except ValidationError as e:
            self.add_error("C", str(e))
            return

        self.variations = []
        for region, variation_names in regions:
            try:
                # TODO(@slamora) translate code to region name
                r = Region.objects.get(name=region)
            except Region.DoesNotExist:
                # if region doesn't exist, maybe it's a location
                try:
                    v = DiatopicVariation.objects.get(name=region)
                    r = v.region
                except DiatopicVariation.DoesNotExist:
                    self.add_error("C", "unkown region '{}'".format(region))
                    r = None

            for name in variation_names:
                try:
                    v = DiatopicVariation.objects.get(name=name)
                except DiatopicVariation.DoesNotExist:
                    self.add_error("C", "unkown location {}".format(name))
                else:
                    if v.region != r:
                        self.add_error("C", "location {} doesn't belong to region {}".format(v, r))
                    else:
                        self.variations.append(v)

        return regions

    def clean_cat(self, value):
        try:
            return split_and_strip(value)
        except TypeError:
            self.add_error("D", self.default_error_messages['required'])

    def clean_es(self, value):
        try:
            return split_and_strip(value)
        except TypeError:
            self.add_error("E", self.default_error_messages['required'])

    def add_error(self, col, message):
        self.errors.append({
            "word": self.term,
            "column": col,
            "message": message,
        })
